import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

#Abre a planilha de clientes que contém Nome, Valor, CPF e Data de Vencimento
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

#Abre o navegador e acessa o site onde será feita a consulta do CPF
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

# Percorre cada linha da planilha a partir da segunda
for linha in pagina_clientes.iter_rows(min_row=2,values_only=True):
    nome, valor, cpf, vencimento = linha

    sleep(4)
    # Localiza o campo de pesquisa do CPF, limpa e insere o CPF do cliente
    campo_pesquisa = driver.find_element(By.XPATH,'//input[@id="cpfInput"]')
    campo_pesquisa.clear()
    sleep(4)
    campo_pesquisa.send_keys(cpf)
    sleep(4)

    # Clica no botão de pesquisa e aguarda o carregamento dos resultados
    pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    pesquisar.click()
    sleep(4)

    # Captura o texto do campo de status para verificar se o cliente está em dia ou pendente
    status = driver.find_element(By.XPATH,'//span[@id="statusLabel"]')

    # Caso o cliente esteja com pagamento em dia, captura a data e método de pagamento
    if status.text == 'em dia': 
        data_pagamento = driver.find_element(By.XPATH,'//p[@id="paymentDate"]')
        data_pagamento_limpo = data_pagamento.text.split()[3]

        metodo_pagamento = driver.find_element(By.XPATH,'//p[@id="paymentMethod"]')
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]
        
        # Abre a planilha de fechamento e registra o cliente com status "em dia"
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'Em dia', data_pagamento_limpo, metodo_pagamento_limpo])
        planilha_fechamento.save('planilha fechamento.xlsx')

    else:
        # Caso o cliente esteja pendente, registra apenas os dados básicos e o status 'Pendente
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome,valor,cpf,vencimento,'Pendente'])

        planilha_fechamento.save('planilha fechamento.xlsx')
